from datetime import date as idate
import re

from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook

from model.matrix_type import MatrixType
from repository.matrix_type import MatrixTypeData
from model.brand import Brand
from repository.brand import BrandData
from model.screen_resolution import ScreenResolution
from repository.screen_resolution import ScreenResolutionData
from model.os import OS
from repository.os import OSData
from repository.currency import CurrencyData
from repository.shop import ShopData
from model.tv import TV
from repository.tv import TVData
from repository.shop_link import ShopLinkData
from model.shop_link import ShopLink
from model.day_price import DayPrice
from repository.day_price import DayPriceData


async def file_upload_handle(
        file, currency_id: int, shop_id: int,
        date: idate, session: AsyncSession):
    try:
        currency = await CurrencyData(session).get_one(currency_id)
        shop = await ShopData(session).get_one(shop_id)
        if not currency or not shop:
            raise Exception('shop or currency not exist')
        wb = load_workbook(file)
        for sheet in wb.worksheets:
            title = sheet.title.lower()
            category: str = None
            if title.find('not') != -1 or title.find('не') != -1:
                category = 'Без Smart TV'

            line_count = 0
            title_dict = {}
            for row in sheet.iter_rows(values_only=True, max_col=70):
                if line_count == 0:
                    for indx, colum_name in enumerate(row):
                        if colum_name is None:
                            break
                        column = str(colum_name).lower()
                        if column.find('описание') != -1:
                            title_dict['description'] = indx
                            continue
                        if column.find('диагональ') != -1:
                            title_dict['diagonal'] = indx
                            continue
                        if column.find('частота обновлени') != -1:
                            title_dict['refresh_rate'] = indx
                            continue
                        if column.find('цвет') != -1 and column.find('рамк') == -1:
                            title_dict['color'] = indx
                            continue
                        if column.find('технология экрана') != -1 or column.find('тип матрицы') != -1:
                            title_dict['matrix_type'] = indx
                            print(indx)
                            continue
                        if column.find('бренд') != -1:
                            title_dict['brand'] = indx
                            continue
                        if column.find('разрешение') != -1 or column.find('стандарт разрешения') != -1:
                            title_dict['screen_resolution'] = indx
                            continue
                        if column.find('операционная система') != -1:
                            title_dict['os'] = indx
                            continue
                        if column.find('цена до скидки') != -1 or column.find('цена со скидкой') != -1:
                            title_dict['full_price'] = indx
                            continue
                        if column.find('цена по карте') != -1 or column.find('цена с картой') != -1:
                            title_dict['card_price'] = indx
                            continue
                        if column.find('цена') != -1:
                            title_dict['price'] = indx
                            continue
                    line_count += 1
                    continue
                break_exit = False

                link: str = None
                name: str = None
                diagonal: int = None
                description: str = None
                refresh_rate: int = None
                color: str = None
                matrix_type: MatrixType = None
                brand: Brand = None
                screen_resolution: ScreenResolution = None
                os: OS = None
                full_price = 0
                card_price = 0
                price = 0
                for indx, cell in enumerate(row):
                    if indx == 0 and cell is None:
                        break_exit = True
                        break
                    if indx == 1:
                        link = str(cell)
                        continue
                    if indx == 2:
                        name = str(cell)
                        name = name.replace('"', '').replace("'", '').replace('`', '').replace('″', '').replace('’', '').replace('”', '').replace('`', '')
                        continue
                    if indx == title_dict.get('description'):
                        description = str(cell)
                        continue
                    if indx == title_dict.get('diagonal'):
                        if cell is None:
                            break
                        cell_str = str(cell).lower()
                        cell_str = re.sub(r'\d+см;', '', cell_str)
                        cell_str = re.sub(r'\d+см', '', cell_str)
                        cell_str = cell_str.replace(' ', '').replace('дюйм', '"').replace('″', '"').replace('\'', '"').replace(';', '"').replace('’', '"').replace('(', '"').replace('”', '"')
                        try:
                            diagonal = int(float(cell_str.split('"')[0]))
                        except Exception as ex:
                            print(ex)
                        continue
                    if indx == title_dict.get('refresh_rate'):
                        if cell is None:
                            break
                        try:
                            refresh_rate = int(str(cell).lower().replace('hz', '').replace('гц', '').replace(' ', ''))
                        except Exception:
                            print('refresh_rate:', str(cell))
                        continue
                    if indx == title_dict.get('color'):
                        color = str(cell)
                        continue
                    if indx == title_dict.get('matrix_type'):
                        if cell is None:
                            break
                        matrix_type_buf = str(cell).lower()
                        if matrix_type_buf.find('tn') != -1:
                            matrix_type = await MatrixTypeData(session).get_by_name('TN')
                        elif matrix_type_buf.find('ips') != -1:
                            matrix_type = await MatrixTypeData(session).get_by_name('IPS')
                        elif matrix_type_buf.find('va') != -1:
                            matrix_type = await MatrixTypeData(session).get_by_name('VA')
                        elif matrix_type_buf.find('qled') != -1:
                            matrix_type = await MatrixTypeData(session).get_by_name('QLED')
                        elif matrix_type_buf.find('dled') != -1 or matrix_type_buf.find('dual led') != -1:
                            matrix_type = await MatrixTypeData(session).get_by_name('DLED')
                        elif matrix_type_buf.find('oled') != -1:
                            matrix_type = await MatrixTypeData(session).get_by_name('OLED')
                        elif matrix_type_buf.find('led') != -1:
                            matrix_type = await MatrixTypeData(session).get_by_name('LED')
                        continue
                    if indx == title_dict.get('brand'):
                        if cell is None:
                            break
                        brand_buf = str(cell).lower().replace('.', '')
                        if brand_buf is None or brand_buf == '':
                            brand = None
                            continue
                        if brand_buf.find('телевизор') != -1:
                            brand = None
                            continue
                        brand = await BrandData(session).get_by_name(brand_buf)
                        continue
                    if indx == title_dict.get('screen_resolution'):
                        if cell is None:
                            break
                        screen_resolution_buf = str(cell).lower().replace(' ', '')
                        if screen_resolution_buf.find('fhd') != -1 or screen_resolution_buf.find('fullhd'):
                            screen_resolution = await ScreenResolutionData(session).get_by_name('Full HD')
                        elif screen_resolution_buf.find('4k') != -1:
                            screen_resolution = await ScreenResolutionData(session).get_by_name('4K')
                        elif screen_resolution_buf.find('2k') != -1:
                            screen_resolution = await ScreenResolutionData(session).get_by_name('2K')
                        elif screen_resolution_buf.find('hd') != -1:
                            screen_resolution = await ScreenResolutionData(session).get_by_name('HD')
                        elif screen_resolution_buf.find('5k') != -1:
                            screen_resolution = await ScreenResolutionData(session).get_by_name('5K')
                        elif screen_resolution_buf.find('8k') != -1:
                            screen_resolution = await ScreenResolutionData(session).get_by_name('8K')
                        continue
                    if indx == title_dict.get('os'):
                        if cell is None:
                            break
                        os_buf = str(cell).lower()
                        if os_buf.find('webos') != -1 or os_buf.find('web os') != -1:
                            os = await OSData(session).get_by_name('Web OS')
                        elif os_buf.find('tizen') != -1:
                            os = await OSData(session).get_by_name('Tizen OS')
                        elif os_buf.find('android') != -1:
                            os = await OSData(session).get_by_name('Android TV')
                        elif os_buf.find('google') != -1:
                            os = await OSData(session).get_by_name('Google TV')
                        elif os_buf.find('yaos') != -1 or os_buf.find('яндекс') != -1:
                            os = await OSData(session).get_by_name('YaOS')
                        elif os_buf.find('салют') != -1:
                            os = await OSData(session).get_by_name('Салют')
                        elif os_buf.find('wildred') != -1:
                            os = await OSData(session).get_by_name('WildRed')
                        elif os_buf.find('vidaa') != -1:
                            os = await OSData(session).get_by_name('VIDAA')
                        elif os_buf.find('без') != -1 or category == 'Без Smart TV':
                            os = await OSData(session).get_by_name('Без Smart TV')
                        continue
                    if indx == title_dict.get('full_price'):
                        try:
                            full_price = float(str(cell).replace(' ', ''))
                        except Exception:
                            print('full_price:', cell)
                        continue
                    if indx == title_dict.get('card_price'):
                        try:
                            card_price = float(str(cell).replace(' ', ''))
                        except Exception:
                            print('card_price:', cell)
                        continue
                    if indx == title_dict.get('price'):
                        try:
                            price = float(str(cell).replace(' ', ''))
                        except Exception:
                            print('price:', cell)
                        continue

                if break_exit:
                    break

                if price == 0 and card_price == 0 and full_price == 0:
                    line_count += 1
                    continue
                tv = await TVData(session).get_by_name(name)
                if not tv:
                    if brand is None:
                        brand = await BrandData(session).get_by_entry(tv_name=name)
                    tv_model = TV(
                        name=name,
                        description=description,
                        diagonal=diagonal,
                        refresh_rate=refresh_rate,
                        color=color[:20] if color else None,
                        matrix_type_id=matrix_type.id if matrix_type else None,
                        brand_id=brand.id if brand else None,
                        screen_resolution_id=screen_resolution.id if screen_resolution else None,
                        os_id=os.id if os else None,
                        )
                    tv = await TVData(session).create_by_model(tv_model)

                shop_link = await ShopLinkData(session).get_by_shop_tv(shop_id, tv.id)

                if not shop_link:
                    shop_link_model = ShopLink(
                        shop_id=shop_id,
                        tv_id=tv.id,
                        link=link,
                        is_active=True
                    )
                    shop_link = await ShopLinkData(session).create_by_model(shop_link_model)

                if card_price != 0:
                    day_price_model = DayPrice(
                            shop_link_id=shop_link.id,
                            price=card_price,
                            name='card_price',
                            currency_id=currency_id,
                            date=date
                        )
                    await DayPriceData(session).create_by_model(day_price_model)
                if full_price != 0:
                    day_price_model = DayPrice(
                            shop_link_id=shop_link.id,
                            price=full_price,
                            name='full_price',
                            currency_id=currency_id,
                            date=date
                        )
                    await DayPriceData(session).create_by_model(day_price_model)
                if price != 0:
                    day_price_model = DayPrice(
                            shop_link_id=shop_link.id,
                            price=price,
                            name='discount_price',
                            currency_id=currency_id,
                            date=date
                        )
                    await DayPriceData(session).create_by_model(day_price_model)

                line_count += 1
    except Exception as e:
        print(e)
        raise e
